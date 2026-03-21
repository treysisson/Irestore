#!/usr/bin/env python3
"""
iRESTORE COO Practice — Demand Planning & S&OP Workbook
Matches the format of iRestore_COO_Practice_v2.xlsx:
  Yellow = calculate / enter a number or formula
  Green  = your written answer
  White  = given data, do not change
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Styles (matching existing workbook) ──────────────────────────────────────

HEADER_FILL = PatternFill(start_color="1B3A5C", end_color="1B3A5C", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
SECTION_FONT = Font(name="Calibri", bold=True, size=11, color="1B3A5C")
BODY_FONT = Font(name="Calibri", size=11)
BOLD_FONT = Font(name="Calibri", bold=True, size=11)
TITLE_FONT = Font(name="Calibri", bold=True, size=13, color="1B3A5C")
NOTE_FONT = Font(name="Calibri", italic=True, size=10, color="666666")
YELLOW = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # calculate
GREEN = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")   # written answer
THIN = Border(
    left=Side(style="thin", color="CCCCCC"), right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"), bottom=Side(style="thin", color="CCCCCC"),
)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", wrap_text=True, vertical="top")

MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

FMT_NUM = '#,##0'
FMT_USD = '$#,##0'
FMT_PCT = '0%'
FMT_PCT1 = '0.0%'
FMT_USD_K = '$#,##0'


def hdr(ws, row, cols, values):
    """Write a header row."""
    for c, v in enumerate(values, cols[0] if isinstance(cols, range) else 1):
        cell = ws.cell(row=row, column=c, value=v)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN


def section_label(ws, row, col, text, max_col=None):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = SECTION_FONT
    if max_col:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=max_col)


def data_cell(ws, row, col, value, fmt=None, font=None, fill=None, align=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font or BODY_FONT
    cell.border = THIN
    cell.alignment = align or (CENTER if col > 1 else WRAP)
    if fmt:
        cell.number_format = fmt
    if fill:
        cell.fill = fill
    return cell


def note_cell(ws, row, col, text):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = NOTE_FONT
    cell.alignment = WRAP
    cell.border = THIN


def yellow(ws, row, col, fmt=None, label="← calculate"):
    """Mark a cell as calculate-here."""
    cell = ws.cell(row=row, column=col)
    cell.fill = YELLOW
    cell.border = THIN
    cell.alignment = CENTER
    if fmt:
        cell.number_format = fmt
    return cell


def green(ws, row, col, label=""):
    """Mark a cell as written-answer."""
    cell = ws.cell(row=row, column=col, value=label)
    cell.fill = GREEN
    cell.border = THIN
    cell.alignment = WRAP
    return cell


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ═══════════════════════════════════════════════════════════════════════════════
# START HERE sheet
# ═══════════════════════════════════════════════════════════════════════════════

def build_start(wb):
    ws = wb.create_sheet("0 - Start Here")
    set_widths(ws, [100])

    ws.cell(row=1, column=1, value="iRESTORE COO INTERVIEW — DEMAND PLANNING PRACTICE").font = Font(name="Calibri", bold=True, size=14, color="1B3A5C")
    ws.cell(row=3, column=1, value="5 scenarios focused on demand planning, S&OP, and inventory — the #1 operational gap Kevin has called out.").font = BODY_FONT
    ws.cell(row=5, column=1, value="  1 - Seasonal Demand Forecast     →  Build a monthly unit plan using 2025 actuals and seasonality; identify Q4 peak risk").font = BODY_FONT
    ws.cell(row=6, column=1, value="  2 - Costco Surge Scenario         →  Model 3x/5x/10x Costco promo volumes; build a contingency plan on a timeline").font = BODY_FONT
    ws.cell(row=7, column=1, value="  3 - Safety Stock & Reorder        →  Calculate reorder points and safety stock given 5-week lead time and demand variability").font = BODY_FONT
    ws.cell(row=8, column=1, value="  4 - New SKU Launch Ramp           →  Forecast demand for a new product with no history; use an analog SKU to calibrate").font = BODY_FONT
    ws.cell(row=9, column=1, value="  5 - Inventory Health & Cash       →  Diagnose overstocked SKUs, calculate trapped cash, and recommend liquidation actions").font = BODY_FONT
    ws.cell(row=10, column=1, value="  6 - Channel P&L                   →  Build a fully-loaded P&L by channel (DTC/Amazon/Costco/B2B); find where you actually make money").font = BODY_FONT
    ws.cell(row=12, column=1, value="COLOR KEY:").font = BOLD_FONT
    ws.cell(row=13, column=1, value="  🟡 Yellow  =  calculate this (enter a number or formula)").font = BODY_FONT
    ws.cell(row=14, column=1, value="  🟢 Green   =  your written answer").font = BODY_FONT
    ws.cell(row=15, column=1, value="  ⬜ White   =  given data, do not change").font = BODY_FONT
    ws.cell(row=17, column=1, value="GRADING NOTE:").font = BOLD_FONT
    ws.cell(row=18, column=1, value="  Show your work — use formulas, not just typed numbers.").font = BODY_FONT
    ws.cell(row=19, column=1, value="  For written answers: lead with the recommendation, then support with numbers.").font = BODY_FONT
    ws.cell(row=20, column=1, value="  Round to the nearest dollar or percentage point — precision is less important than logic.").font = BODY_FONT

    ws.sheet_properties.tabColor = "1B3A5C"


# ═══════════════════════════════════════════════════════════════════════════════
# SCENARIO 1: Seasonal Demand Forecast
# ═══════════════════════════════════════════════════════════════════════════════

def build_scenario_1(wb):
    ws = wb.create_sheet("1 - Seasonal Demand")
    set_widths(ws, [34] + [12] * 13 + [14])

    ws.cell(row=1, column=1, value="SCENARIO 1: SEASONAL DEMAND FORECAST  |  Build the 2026 Monthly Unit Plan").font = TITLE_FONT

    # SECTION A — 2025 Actuals
    r = 3
    section_label(ws, r, 1, "SECTION A — 2025 Actual Units Sold by SKU (given data)", 14)
    r += 1
    headers = ["SKU"] + MONTHS + ["FY 2025"]
    hdr(ws, r, range(1, len(headers) + 1), headers)
    r += 1

    # Realistic monthly unit data per SKU (sums to annual totals)
    sku_data = {
        "Elite Helmet":         [120, 115, 140, 155, 160, 170, 175, 170, 190, 215, 310, 280],
        "Professional Helmet":  [160, 155, 185, 200, 210, 225, 230, 225, 255, 290, 420, 370],
        "Essential Helmet":     [140, 135, 160, 175, 180, 195, 200, 195, 220, 255, 380, 340],
        "Illumina Face Mask":   [85, 80, 100, 110, 115, 125, 130, 125, 145, 170, 260, 230],
        "Illumina Neck & Chest":[50, 45, 55, 60, 65, 70, 72, 70, 80, 95, 150, 130],
        "Serums & Supplements": [850, 820, 980, 1080, 1130, 1220, 1260, 1220, 1390, 1600, 2400, 2100],
    }
    data_start = r
    for sku, monthly in sku_data.items():
        data_cell(ws, r, 1, sku, font=BOLD_FONT)
        for c, v in enumerate(monthly, 2):
            data_cell(ws, r, c, v, fmt=FMT_NUM)
        data_cell(ws, r, 14, sum(monthly), fmt=FMT_NUM, font=BOLD_FONT)
        r += 1
    # Total row
    data_cell(ws, r, 1, "TOTAL", font=BOLD_FONT)
    for c in range(2, 15):
        col_l = get_column_letter(c)
        data_cell(ws, r, c, f"=SUM({col_l}{data_start}:{col_l}{r-1})", fmt=FMT_NUM, font=BOLD_FONT)
    r += 1

    # Note
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=14)
    note_cell(ws, r, 1, "* iRESTORE is ~$90M annual revenue. 80% YoY growth targeted for 2026. Q4 = ~39% of annual volume. Costco skews even more heavily to Q4.")
    r += 2

    # SECTION B — Reference: ASP and margins
    section_label(ws, r, 1, "SECTION B — Reference: ASP & Channel Mix (given)", 14)
    r += 1
    ref_headers = ["SKU", "ASP (DTC)", "ASP (Amazon)", "DTC %", "Amazon %", "Costco %", "B2B %", "Return Rate", "Gross Margin"]
    hdr(ws, r, range(1, len(ref_headers) + 1), ref_headers)
    r += 1

    ref_data = [
        ("Elite Helmet", 1899, 1699, 0.55, 0.25, 0.15, 0.05, 0.05, 0.74),
        ("Professional Helmet", 899, 799, 0.50, 0.35, 0.10, 0.05, 0.08, 0.70),
        ("Essential Helmet", 499, 449, 0.35, 0.40, 0.20, 0.05, 0.12, 0.64),
        ("Illumina Face Mask", 499, 449, 0.60, 0.30, 0.05, 0.05, 0.09, 0.66),
        ("Illumina Neck & Chest", 399, 359, 0.55, 0.35, 0.05, 0.05, 0.14, 0.62),
        ("Serums & Supplements", 55, 49, 0.65, 0.30, 0.00, 0.05, 0.02, 0.82),
    ]
    for row_d in ref_data:
        data_cell(ws, r, 1, row_d[0], font=BOLD_FONT)
        data_cell(ws, r, 2, row_d[1], fmt=FMT_USD)
        data_cell(ws, r, 3, row_d[2], fmt=FMT_USD)
        for c in range(4, 9):
            data_cell(ws, r, c, row_d[c - 1], fmt=FMT_PCT)
        data_cell(ws, r, 9, row_d[8], fmt=FMT_PCT)
        r += 1
    r += 1

    # SECTION C — Calculate seasonality index
    section_label(ws, r, 1, "SECTION C — Calculate Seasonality Index  (yellow = your work)", 14)
    r += 1
    si_headers = ["Month"] + MONTHS
    hdr(ws, r, range(1, len(si_headers) + 1), si_headers)
    r += 1

    data_cell(ws, r, 1, "Total 2025 Units (from Section A)", font=BOLD_FONT)
    for c in range(2, 14):
        yellow(ws, r, c, fmt=FMT_NUM)
    r += 1
    data_cell(ws, r, 1, "Seasonality Index\n(= month units ÷ annual total)", font=BOLD_FONT)
    for c in range(2, 14):
        yellow(ws, r, c, fmt=FMT_PCT1)
    r += 1
    data_cell(ws, r, 1, "Which months are above 1/12 (8.3%)?", font=BOLD_FONT)
    green(ws, r, 2)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=13)
    r += 2

    # SECTION D — Build 2026 forecast
    section_label(ws, r, 1, "SECTION D — Your 2026 Demand Forecast  (yellow = your work)", 14)
    r += 1
    note_cell(ws, r, 1, "Using 80% YoY growth × your seasonality index, build the monthly unit plan for each SKU. Then answer the questions below.")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=14)
    r += 1
    fc_headers = ["SKU"] + MONTHS + ["FY 2026"]
    hdr(ws, r, range(1, len(fc_headers) + 1), fc_headers)
    r += 1
    fc_start = r
    for sku in sku_data.keys():
        data_cell(ws, r, 1, sku, font=BOLD_FONT)
        for c in range(2, 14):
            yellow(ws, r, c, fmt=FMT_NUM)
        yellow(ws, r, 14, fmt=FMT_NUM)  # total
        r += 1
    data_cell(ws, r, 1, "TOTAL", font=BOLD_FONT)
    for c in range(2, 15):
        col_l = get_column_letter(c)
        data_cell(ws, r, c, f"=SUM({col_l}{fc_start}:{col_l}{r-1})", fmt=FMT_NUM, font=BOLD_FONT)
        ws.cell(row=r, column=c).fill = YELLOW
    r += 1

    # Revenue check
    data_cell(ws, r, 1, "Revenue Check: FY 2026 Total Revenue", font=BOLD_FONT)
    yellow(ws, r, 2, fmt=FMT_USD)
    note_cell(ws, r, 3, "Should be in the ~$155-162M range at 80% growth")
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=14)
    r += 2

    # SECTION E — Questions
    section_label(ws, r, 1, "SECTION E — Questions  (green = written answer)", 14)
    r += 1
    questions = [
        ("Q1: In which month does total unit demand first exceed 2x the January level? What does that tell you about when to start surge planning?", None),
        ("Q2: If the warehouse can handle 800 DTC orders/day, on which months does your forecast exceed that? What's your contingency?", None),
        ("Q3: Given 5-week ocean lead time, what's the LAST month you can place a production order and receive it before November peak?", "Hint: count backwards from Nov 1."),
        ("Q4: Costco's Q4 share is even more concentrated than DTC. If Costco does a Nov endcap promo, how many Professional Helmet units should you pre-stage by Oct 15?", None),
        ("Q5: Kevin wants to grow Serums & Supplements into a subscription business. How does that change the demand curve shape vs. device SKUs?", None),
    ]
    for q, hint in questions:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=14)
        data_cell(ws, r, 1, q, font=BODY_FONT)
        ws.cell(row=r, column=1).alignment = WRAP
        r += 1
        if hint:
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=14)
            note_cell(ws, r, 1, hint)
            r += 1
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=14)
        green(ws, r, 1)
        ws.row_dimensions[r].height = 50
        r += 1

    ws.sheet_properties.tabColor = "1B3A5C"


# ═══════════════════════════════════════════════════════════════════════════════
# SCENARIO 2: Costco Surge — What Would You Have Done?
# ═══════════════════════════════════════════════════════════════════════════════

def build_scenario_2(wb):
    ws = wb.create_sheet("2 - Costco Surge")
    set_widths(ws, [36] + [14] * 5 + [20])

    ws.cell(row=1, column=1, value="SCENARIO 2: COSTCO Q4 SURGE  |  Plan for 3x, 5x, or 10x — What Would You Have Done?").font = TITLE_FONT

    r = 3
    section_label(ws, r, 1, "SECTION A — Current State as of August (given data)", 7)
    r += 1
    hdr(ws, r, range(1, 4), ["Metric", "Value", "Notes"])
    r += 1

    state = [
        ("Professional Helmet — on hand", "8,200 units", "La Mirada warehouse"),
        ("Professional Helmet — in transit", "3,000 units", "Arriving ~Sep 5 via ocean"),
        ("Factory max output / month", "5,000 units", "Wellmike, Shenzhen; 7,000 w/ overtime"),
        ("La Mirada daily pick/pack capacity", "800 orders/day", "Current staffing, single shift"),
        ("La Mirada max capacity (2 shifts + temps)", "1,400 orders/day", "Need 3-week lead time for temp labor"),
        ("Ocean freight lead time (door-to-door)", "5 weeks", "Yantian → LA"),
        ("Air freight lead time", "5-7 days", "~$18/unit vs. $6/unit ocean"),
        ("DTC avg daily orders (all SKUs, non-peak)", "450/day", "Mix of all SKUs"),
        ("Amazon avg daily orders (non-peak)", "280/day", "Mix of all SKUs"),
        ("Costco normal monthly PO", "600-800 units", "Professional Helmet"),
        ("Last year's Costco November promo", "1,200 units / 2 weeks", "Costco buyer says 'expect 3-5x'"),
    ]
    for metric, val, notes in state:
        data_cell(ws, r, 1, metric)
        data_cell(ws, r, 2, val, font=BOLD_FONT)
        note_cell(ws, r, 3, notes)
        r += 1
    r += 1

    # SECTION B — Scenario modeling
    section_label(ws, r, 1, "SECTION B — Model Each Scenario  (yellow = your work)", 7)
    r += 1
    hdr(ws, r, range(1, 6), ["Metric", "Base (3x)", "Likely (5x)", "Hot (8x)", "Actual (10x)"])
    r += 1

    rows_b = [
        ("Costco promo units (2 weeks)", 3600, 6000, 9600, 12000),
        ("Total Nov daily orders (all channels, estimated)", None, None, None, None),
        ("Warehouse capacity gap vs. 800/day", None, None, None, None),
        ("Warehouse capacity gap vs. 1,400/day (w/ temps)", None, None, None, None),
        ("Additional Professional Helmet units needed\n(vs. 8,200 on hand + 3,000 in transit)", None, None, None, None),
        ("Last ocean order date to arrive by Oct 25", None, None, None, None),
        ("Air freight units if ocean misses", None, None, None, None),
        ("Air freight cost premium over ocean", None, None, None, None),
        ("Temp labor cost (3 weeks setup + Nov)", None, None, None, None),
        ("DTC fulfillment delay (estimated days)", None, None, None, None),
        ("Total incremental cost to handle this scenario", None, None, None, None),
    ]
    for label, *vals in rows_b:
        data_cell(ws, r, 1, label, font=BOLD_FONT)
        for c, v in enumerate(vals, 2):
            if v is not None:
                data_cell(ws, r, c, v, fmt=FMT_NUM)
            else:
                yellow(ws, r, c, fmt=FMT_NUM)
        r += 1
    r += 1

    # SECTION C — Action timeline
    section_label(ws, r, 1, "SECTION C — Your Action Plan  (yellow = actions, green = written answer)", 7)
    r += 1
    hdr(ws, r, range(1, 6), ["Week Of", "Action You'd Take", "Owner", "Est. Cost", "What Breaks If You Don't"])
    r += 1
    weeks = [
        "Aug 11 (Today — Costco confirms promo)",
        "Aug 18 (Place production order?)",
        "Aug 25",
        "Sep 1 (In-transit shipment arrives)",
        "Sep 15 (Hire temp labor decision)",
        "Oct 1 (Last ocean order for Nov arrival)",
        "Oct 15 (Pre-stage Costco inventory)",
        "Nov 1 (Promo starts — all hands on deck)",
        "Nov 15 (Promo week 2)",
        "Nov 24 (Black Friday)",
        "Dec 1 (Recovery / catch up DTC backlog)",
    ]
    for wk in weeks:
        data_cell(ws, r, 1, wk, font=BOLD_FONT)
        for c in range(2, 6):
            yellow(ws, r, c)
        r += 1
    r += 1

    # SECTION D — Questions
    section_label(ws, r, 1, "SECTION D — Questions  (green = written answer)", 7)
    r += 1
    questions = [
        "Q1: Brandon's response to the actual Q4 crisis was 'We can't do anything.' Name 3 specific things you'd have had in place by September that prevent this answer.",
        "Q2: At what Costco volume level does air freight become the rational choice despite 3x cost? Show the math — what's the breakeven?",
        "Q3: How do you protect DTC fulfillment SLAs (ship within 2 days) when Costco bulk prep is consuming warehouse capacity?",
        "Q4: The Costco buyer says 'expect 3-5x.' You think 10x is possible. How do you plan for 10x without committing capital to 10x? What's the optionality play?",
        "Q5: What S&OP changes would you implement so this scenario NEVER catches the company off guard again?",
    ]
    for q in questions:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        data_cell(ws, r, 1, q, font=BODY_FONT)
        ws.cell(row=r, column=1).alignment = WRAP
        r += 1
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        green(ws, r, 1)
        ws.row_dimensions[r].height = 55
        r += 1

    ws.sheet_properties.tabColor = "C0392B"


# ═══════════════════════════════════════════════════════════════════════════════
# SCENARIO 3: Safety Stock & Reorder Point Calculator
# ═══════════════════════════════════════════════════════════════════════════════

def build_scenario_3(wb):
    ws = wb.create_sheet("3 - Safety Stock & Reorder")
    set_widths(ws, [34] + [15] * 7 + [18])

    ws.cell(row=1, column=1, value="SCENARIO 3: SAFETY STOCK & REORDER POINTS  |  5-Week Lead Time from Shenzhen").font = TITLE_FONT

    r = 3
    section_label(ws, r, 1, "SECTION A — Weekly Demand Data by SKU  (given: last 12 weeks)", 9)
    r += 1
    hdr(ws, r, range(1, 14), ["SKU"] + [f"Wk {i}" for i in range(1, 13)])
    r += 1

    # Weekly demand data with realistic variability
    weekly = {
        "Elite Helmet":         [38, 42, 35, 44, 40, 37, 51, 39, 43, 46, 55, 48],
        "Professional Helmet":  [52, 48, 55, 61, 50, 58, 63, 54, 59, 66, 72, 60],
        "Essential Helmet":     [45, 50, 42, 48, 53, 46, 44, 55, 49, 52, 58, 47],
        "Illumina Face Mask":   [32, 28, 35, 30, 33, 36, 31, 29, 34, 38, 42, 35],
        "Illumina Neck & Chest":[18, 15, 20, 17, 19, 16, 22, 18, 15, 21, 24, 19],
        "Serums & Supplements": [310, 290, 330, 340, 305, 325, 350, 315, 335, 360, 395, 345],
    }
    wk_start = r
    for sku, wks in weekly.items():
        data_cell(ws, r, 1, sku, font=BOLD_FONT)
        for c, v in enumerate(wks, 2):
            data_cell(ws, r, c, v, fmt=FMT_NUM)
        r += 1
    r += 1

    # SECTION B — Calculate stats
    section_label(ws, r, 1, "SECTION B — Calculate Demand Statistics  (yellow = your work)", 9)
    r += 1
    hdr(ws, r, range(1, 9), ["SKU", "Avg Weekly\nDemand", "Std Dev\n(weekly)", "Coefficient of\nVariation", "Max Week", "Min Week", "Max/Avg\nRatio", "Notes"])
    r += 1
    stat_start = r
    for sku in weekly.keys():
        data_cell(ws, r, 1, sku, font=BOLD_FONT)
        for c in range(2, 8):
            yellow(ws, r, c)
        green(ws, r, 8)
        r += 1
    r += 1
    note_cell(ws, r, 1, "* Coefficient of Variation = Std Dev ÷ Avg. Higher CV = more variable demand = more safety stock needed.")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    r += 2

    # SECTION C — Reorder point calculation
    section_label(ws, r, 1, "SECTION C — Reorder Point & Safety Stock  (yellow = your work)", 9)
    r += 1
    note_cell(ws, r, 1, "Lead time = 5 weeks (ocean). Service level target = 95% (z = 1.65). Safety Stock = z × σ × √(lead time in weeks)")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    r += 1
    hdr(ws, r, range(1, 9), ["SKU", "Lead Time\n(weeks)", "Safety Stock\n(units)", "Reorder Point\n(units)", "Current\nOn Hand", "Weeks of\nSupply", "Status\n(OK / Reorder / Urgent)", "Order Qty\nRecommended"])
    r += 1
    on_hand = {
        "Elite Helmet": 420,
        "Professional Helmet": 580,
        "Essential Helmet": 510,
        "Illumina Face Mask": 340,
        "Illumina Neck & Chest": 620,
        "Serums & Supplements": 4200,
    }
    for sku in weekly.keys():
        data_cell(ws, r, 1, sku, font=BOLD_FONT)
        data_cell(ws, r, 2, 5, fmt=FMT_NUM)
        yellow(ws, r, 3, fmt=FMT_NUM)  # safety stock
        yellow(ws, r, 4, fmt=FMT_NUM)  # reorder point
        data_cell(ws, r, 5, on_hand[sku], fmt=FMT_NUM, font=BOLD_FONT)
        yellow(ws, r, 6, fmt='0.0')  # weeks of supply
        yellow(ws, r, 7)  # status
        yellow(ws, r, 8, fmt=FMT_NUM)  # order qty
        r += 1
    r += 1

    # SECTION D — Scenario: What if lead time doubles?
    section_label(ws, r, 1, "SECTION D — Stress Test: What If Lead Time Doubles to 10 Weeks?  (yellow = your work)", 9)
    r += 1
    note_cell(ws, r, 1, "Port congestion, factory delay, or Chinese New Year could push lead time to 10 weeks. Recalculate safety stock and reorder points.")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    r += 1
    hdr(ws, r, range(1, 7), ["SKU", "New Safety\nStock (10wk LT)", "New Reorder\nPoint", "Additional Inventory\n$ Required", "COGS/Unit", "Incremental\nCash Tied Up"])
    r += 1
    cogs = {"Elite Helmet": 494, "Professional Helmet": 270, "Essential Helmet": 180,
            "Illumina Face Mask": 170, "Illumina Neck & Chest": 152, "Serums & Supplements": 10}
    for sku in weekly.keys():
        data_cell(ws, r, 1, sku, font=BOLD_FONT)
        yellow(ws, r, 2, fmt=FMT_NUM)
        yellow(ws, r, 3, fmt=FMT_NUM)
        yellow(ws, r, 4, fmt=FMT_NUM)
        data_cell(ws, r, 5, cogs[sku], fmt=FMT_USD)
        yellow(ws, r, 6, fmt=FMT_USD)
    r += 1
    # Total
    data_cell(ws, r, 1, "TOTAL INCREMENTAL CASH", font=BOLD_FONT)
    yellow(ws, r, 6, fmt=FMT_USD)
    r += 2

    # Questions
    section_label(ws, r, 1, "SECTION E — Questions  (green = written answer)", 9)
    r += 1
    questions = [
        "Q1: Which SKU has the most variable demand (highest CV)? What does that imply about how you should order it differently vs. a stable SKU?",
        "Q2: The Neck & Chest panel has 620 units on hand but only sells ~18/week. That's 34 weeks of supply. Is that a problem? What would you do?",
        "Q3: If you could only afford to hold safety stock on 3 of the 6 SKUs, which 3 would you choose and why?",
        "Q4: Chinese New Year shuts the factory for 3 weeks in January. How does that change your October/November ordering?",
        "Q5: Kevin's father runs the factory. Does that change your safety stock strategy vs. a third-party contract manufacturer? How?",
    ]
    for q in questions:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        data_cell(ws, r, 1, q, font=BODY_FONT)
        ws.cell(row=r, column=1).alignment = WRAP
        r += 1
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        green(ws, r, 1)
        ws.row_dimensions[r].height = 55
        r += 1

    ws.sheet_properties.tabColor = "27AE60"


# ═══════════════════════════════════════════════════════════════════════════════
# SCENARIO 4: New SKU Launch Ramp (Body Panel)
# ═══════════════════════════════════════════════════════════════════════════════

def build_scenario_4(wb):
    ws = wb.create_sheet("4 - New SKU Launch Ramp")
    set_widths(ws, [32] + [13] * 12 + [14])

    ws.cell(row=1, column=1, value="SCENARIO 4: NEW SKU LAUNCH  |  Body Panel ($1,295) Demand Ramp — Zero History").font = TITLE_FONT

    r = 3
    section_label(ws, r, 1, "SECTION A — Analog: Illumina Face Mask Launch (first 12 months, given)", 14)
    r += 1
    hdr(ws, r, range(1, 14), ["Metric"] + MONTHS)
    r += 1

    analog = {
        "Units Sold":       [320, 480, 710, 890, 1050, 1180, 1280, 1350, 1520, 1680, 2400, 2340],
        "Return Rate":      [0.15, 0.13, 0.12, 0.11, 0.10, 0.09, 0.09, 0.09, 0.08, 0.08, 0.10, 0.09],
        "Marketing Spend ($K)": [180, 150, 140, 130, 120, 110, 100, 95, 110, 120, 200, 160],
        "CAC ($)":          [165, 142, 128, 115, 98, 82, 74, 68, 72, 71, 83, 68],
    }
    for label, vals in analog.items():
        data_cell(ws, r, 1, label, font=BOLD_FONT)
        fmt = FMT_PCT if "Rate" in label else (FMT_USD if "$" in label else FMT_NUM)
        for c, v in enumerate(vals, 2):
            data_cell(ws, r, c, v, fmt=fmt)
        r += 1
    r += 1

    # Face Mask context
    note_cell(ws, r, 1, "Face Mask: $499 ASP, $170 COGS, 66% margin. Launched to existing customer base with email + influencer push. Steady state reached ~month 6.")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=13)
    r += 2

    # SECTION B — Body Panel givens
    section_label(ws, r, 1, "SECTION B — Body Panel Product Data (given)", 14)
    r += 1
    givens = [
        ("ASP (DTC)", "$1,295"),
        ("ASP (Amazon)", "$1,195"),
        ("COGS / unit", "$389"),
        ("Gross Margin", "70%"),
        ("Launch date", "April 2026"),
        ("Initial production run (MOQ)", "3,000 units"),
        ("Reorder MOQ", "1,500 units"),
        ("Initial tooling lead time", "6 weeks"),
        ("Reorder lead time", "4 weeks"),
        ("Launch marketing budget (months 1-3)", "$450K total"),
        ("Ongoing marketing budget", "$80-120K/month"),
        ("Pre-launch waitlist signups", "2,800"),
        ("Kevin's Year 1 target", "15,000 units"),
        ("Marketing's Year 1 estimate", "25,000 units"),
        ("Supply chain's Year 1 estimate", "8,000 units"),
    ]
    for label, val in givens:
        data_cell(ws, r, 1, label)
        data_cell(ws, r, 2, val, font=BOLD_FONT)
        r += 1
    r += 1

    # SECTION C — Your forecast
    section_label(ws, r, 1, "SECTION C — Your Body Panel Demand Forecast: Apr 2026 – Mar 2027  (yellow = your work)", 14)
    r += 1
    ramp_months = ["Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "Jan'27", "Feb'27", "Mar'27", "FY Total"]
    hdr(ws, r, range(1, len(ramp_months) + 2), ["Metric"] + ramp_months)
    r += 1

    ramp_rows = [
        "Units — Base Case",
        "Units — Upside Case",
        "Units — Downside Case",
        "Production Orders Placed (units)",
        "Ending Inventory (EOM)",
        "Weeks of Supply",
        "Marketing Spend ($K)",
        "Estimated CAC ($)",
    ]
    for label in ramp_rows:
        data_cell(ws, r, 1, label, font=BOLD_FONT)
        for c in range(2, 15):
            yellow(ws, r, c, fmt=FMT_NUM)
        r += 1
    r += 1

    # SECTION D — Unit economics check
    section_label(ws, r, 1, "SECTION D — Contribution Margin Check  (yellow = your work)", 14)
    r += 1
    hdr(ws, r, range(1, 5), ["Metric", "Base Case", "Upside", "Downside"])
    r += 1
    cm_rows = [
        "Year 1 Gross Revenue",
        "Year 1 COGS",
        "Year 1 Gross Profit",
        "Year 1 Marketing Spend",
        "Year 1 Contribution Margin $",
        "Year 1 Contribution Margin %",
        "Breakeven Unit Volume",
        "Months to Breakeven",
    ]
    for label in cm_rows:
        data_cell(ws, r, 1, label, font=BOLD_FONT)
        for c in range(2, 5):
            yellow(ws, r, c, fmt=FMT_USD if "%" not in label else FMT_PCT)
        r += 1
    r += 1

    # Questions
    section_label(ws, r, 1, "SECTION E — Questions  (green = written answer)", 14)
    r += 1
    questions = [
        "Q1: The Face Mask ($499) reached steady state in ~6 months. The Body Panel is 2.6x the price. How does that change the ramp shape? Faster or slower? Why?",
        "Q2: You have 2,800 waitlist signups. What Month 1 conversion rate do you assume, and how does that inform your base vs. upside case?",
        "Q3: Kevin says 15K. Marketing says 25K. Supply chain says 8K. What's your number and how did you get there?",
        "Q4: At what point in the ramp would you place your first reorder? What signal triggers it — a date, a unit threshold, or something else?",
        "Q5: If Month 1-2 sales come in 40% below your base case, what's your playbook? When do you cut spend vs. ride it out?",
    ]
    for q in questions:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=14)
        data_cell(ws, r, 1, q, font=BODY_FONT)
        ws.cell(row=r, column=1).alignment = WRAP
        r += 1
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=14)
        green(ws, r, 1)
        ws.row_dimensions[r].height = 55
        r += 1

    ws.sheet_properties.tabColor = "8E44AD"


# ═══════════════════════════════════════════════════════════════════════════════
# SCENARIO 5: Inventory Health & Cash Trapped
# ═══════════════════════════════════════════════════════════════════════════════

def build_scenario_5(wb):
    ws = wb.create_sheet("5 - Inventory Health")
    set_widths(ws, [32] + [14] * 8)

    ws.cell(row=1, column=1, value="SCENARIO 5: INVENTORY HEALTH  |  Where's the Cash Trapped?").font = TITLE_FONT

    r = 3
    section_label(ws, r, 1, "SECTION A — Inventory Snapshot (given: as of March 1, 2026)", 9)
    r += 1
    hdr(ws, r, range(1, 9), ["SKU", "Units\nOn Hand", "COGS/\nUnit", "Avg Monthly\nUnits Sold", "Inventory\nValue ($)", "Days of\nSupply", "Months of\nSupply", "Notes"])
    r += 1

    inv = [
        ("Elite Helmet", 4200, 494, 1500, None, None, None, "Franchise SKU — premium margin"),
        ("Professional Helmet", 6800, 270, 2083, None, None, None, "Volume workhorse — Costco + Amazon"),
        ("Essential Helmet", 8500, 180, 1800, None, None, None, "Entry level — highest return rate"),
        ("Illumina Face Mask", 3200, 170, 1350, None, None, None, "Growing — new category"),
        ("Illumina Neck & Chest", 5100, 152, 750, None, None, None, "Slow seller, 14% returns"),
        ("Illumina Eye Mask", 2400, 100, 600, None, None, None, "Small SKU"),
        ("Serums & Supplements", 42000, 10, 13500, None, None, None, "High volume, high margin, low COGS"),
    ]
    inv_start = r
    for sku, units, cogs_u, monthly, _, _, _, notes in inv:
        data_cell(ws, r, 1, sku, font=BOLD_FONT)
        data_cell(ws, r, 2, units, fmt=FMT_NUM)
        data_cell(ws, r, 3, cogs_u, fmt=FMT_USD)
        data_cell(ws, r, 4, monthly, fmt=FMT_NUM)
        yellow(ws, r, 5, fmt=FMT_USD)  # inv value = units × cogs
        yellow(ws, r, 6, fmt=FMT_NUM)  # days = (units/monthly)*30
        yellow(ws, r, 7, fmt='0.0')    # months = units/monthly
        note_cell(ws, r, 8, notes)
        r += 1
    data_cell(ws, r, 1, "TOTAL", font=BOLD_FONT)
    yellow(ws, r, 5, fmt=FMT_USD)
    r += 2

    # SECTION B — Aging
    section_label(ws, r, 1, "SECTION B — Inventory Aging (given: units by age bucket)", 9)
    r += 1
    hdr(ws, r, range(1, 8), ["SKU", "0-30\ndays", "31-60\ndays", "61-90\ndays", "91-120\ndays", "120+\ndays", "% Over\n90 Days"])
    r += 1

    aging = [
        ("Elite Helmet", 1500, 1200, 800, 400, 300),
        ("Professional Helmet", 2800, 2000, 1200, 500, 300),
        ("Essential Helmet", 2200, 1800, 1500, 1400, 1600),
        ("Illumina Face Mask", 1400, 1000, 500, 200, 100),
        ("Illumina Neck & Chest", 800, 900, 1000, 1200, 1200),
        ("Illumina Eye Mask", 800, 700, 500, 250, 150),
        ("Serums & Supplements", 18000, 12000, 8000, 3000, 1000),
    ]
    for sku, *buckets in aging:
        data_cell(ws, r, 1, sku, font=BOLD_FONT)
        for c, v in enumerate(buckets, 2):
            data_cell(ws, r, c, v, fmt=FMT_NUM)
        yellow(ws, r, 7, fmt=FMT_PCT1)  # % over 90
        r += 1
    r += 1

    # SECTION C — Diagnose
    section_label(ws, r, 1, "SECTION C — Your Diagnosis  (yellow = calculations, green = written answers)", 9)
    r += 1
    hdr(ws, r, range(1, 8), ["SKU", "Health\nRating\n(A/B/C/D)", "Cash\nTrapped ($)", "Root Cause\n(demand? forecast?\nproduct?)", "Recommended\nAction", "Units\nAffected", "Cash\nFreed ($)"])
    r += 1
    for sku, *_ in aging:
        data_cell(ws, r, 1, sku, font=BOLD_FONT)
        yellow(ws, r, 2)
        yellow(ws, r, 3, fmt=FMT_USD)
        green(ws, r, 4)
        green(ws, r, 5)
        yellow(ws, r, 6, fmt=FMT_NUM)
        yellow(ws, r, 7, fmt=FMT_USD)
        r += 1
    data_cell(ws, r, 1, "TOTAL CASH FREED", font=BOLD_FONT)
    yellow(ws, r, 7, fmt=FMT_USD)
    r += 2

    # Questions
    section_label(ws, r, 1, "SECTION D — Questions  (green = written answer)", 9)
    r += 1
    questions = [
        "Q1: Which 2 SKUs have the worst inventory health? What specifically would you do about each within 30 days?",
        "Q2: The Neck & Chest has 6.8 months of supply and 47% of units over 90 days old. Is this a demand problem, a forecast problem, or a product problem? How do you diagnose?",
        "Q3: Kevin wants 80% growth, which means buying MORE inventory. Finance wants to free up cash. How do you do both? Isn't that contradictory?",
        "Q4: Essential Helmet has the worst aging but it's also the entry-level product with the highest Amazon search volume. Do you discount, liquidate, or hold? What's the cost of each option?",
        "Q5: What 3 KPIs would you put on a weekly inventory dashboard? Who sees it? What triggers an escalation?",
    ]
    for q in questions:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        data_cell(ws, r, 1, q, font=BODY_FONT)
        ws.cell(row=r, column=1).alignment = WRAP
        r += 1
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        green(ws, r, 1)
        ws.row_dimensions[r].height = 55
        r += 1

    ws.sheet_properties.tabColor = "F39C12"


# ═══════════════════════════════════════════════════════════════════════════════
# SCENARIO 6: Channel P&L — Where Do You Actually Make Money?
# ═══════════════════════════════════════════════════════════════════════════════

def build_scenario_6(wb):
    ws = wb.create_sheet("6 - Channel P&L")
    set_widths(ws, [36] + [15] * 5 + [18])

    ws.cell(row=1, column=1, value="SCENARIO 6: CHANNEL P&L  |  DTC vs. Amazon vs. Costco vs. B2B — Where Do You Actually Make Money?").font = TITLE_FONT

    r = 3
    section_label(ws, r, 1, "SECTION A — Revenue & Volume by Channel  (given: FY 2025)", 7)
    r += 1
    hdr(ws, r, range(1, 7), ["Metric", "DTC", "Amazon", "Costco", "B2B / Clinic", "Total"])
    r += 1

    # Given data rows
    given_a = [
        ("Gross Units Sold", 48000, 38000, 12000, 4500, None),
        ("Blended ASP", 682, 542, 449, 895, None),
        ("Gross Revenue", None, None, None, None, None),
        ("Return Rate", 0.085, 0.15, 0.28, 0.03, None),
        ("Units Returned", None, None, None, None, None),
        ("Net Units Sold", None, None, None, None, None),
        ("Net Revenue", None, None, None, None, None),
    ]
    for label, *vals in given_a:
        data_cell(ws, r, 1, label, font=BOLD_FONT)
        is_given = label in ("Gross Units Sold", "Blended ASP", "Return Rate")
        for c, v in enumerate(vals, 2):
            if is_given and v is not None:
                fmt = FMT_PCT if "Rate" in label else (FMT_USD if "ASP" in label else FMT_NUM)
                data_cell(ws, r, c, v, fmt=fmt)
            else:
                fmt = FMT_PCT if "Rate" in label else (FMT_USD if "Revenue" in label or "ASP" in label else FMT_NUM)
                yellow(ws, r, c, fmt=fmt)
        r += 1
    r += 1

    # SECTION B — Cost structure
    section_label(ws, r, 1, "SECTION B — Cost Structure by Channel  (given)", 7)
    r += 1
    hdr(ws, r, range(1, 7), ["Cost Line", "DTC", "Amazon", "Costco", "B2B / Clinic", "Notes"])
    r += 1

    costs = [
        ("COGS % of Net Revenue", 0.32, 0.32, 0.32, 0.32, "Same product, same factory cost"),
        ("Fulfillment / Shipping per unit", 8.20, 0, 3.10, 0, "Amazon = FBA (in fees); B2B = FOB"),
        ("Platform / Marketplace Fees %", 0, 0.15, 0, 0, "Amazon referral + FBA fee"),
        ("Wholesale Discount / Margin %", 0, 0, 0.35, 0.25, "Costco takes 35%; clinics negotiate 25%"),
        ("Return Processing Cost / unit", 21.60, 11.80, 18.30, 0, "Outbound + return ship + inspection"),
        ("Ad Spend % of Gross Revenue", 0.22, 0.18, 0.03, 0.05, "DTC heaviest; Costco = coop ad fund"),
        ("Refurb Recovery Rate", 0.55, 0.40, 0.25, 0.75, "% of returns resold as refurbished"),
        ("Refurb Net Recovery / unit", 71, 54, 47, 183, "Resale value minus refurb processing"),
    ]
    for label, *vals in costs:
        data_cell(ws, r, 1, label, font=BOLD_FONT)
        for c, v in enumerate(vals[:4], 2):
            fmt = FMT_PCT if "%" in label else FMT_USD
            data_cell(ws, r, c, v, fmt=fmt)
        note_cell(ws, r, 6, vals[4])
        r += 1
    r += 1

    # SECTION C — Build the P&L
    section_label(ws, r, 1, "SECTION C — Build the Channel P&L  (yellow = your work)", 7)
    r += 1
    hdr(ws, r, range(1, 7), ["P&L Line", "DTC", "Amazon", "Costco", "B2B / Clinic", "Total"])
    r += 1

    pl_rows = [
        "Net Revenue",
        "COGS",
        "Gross Profit",
        "Gross Margin %",
        "",
        "Fulfillment / Shipping",
        "Platform / Marketplace Fees",
        "Wholesale Discount Given",
        "Ad Spend",
        "Return Processing Costs (total)",
        "Refurb Recovery (credit)",
        "Net Return Cost",
        "",
        "Contribution Profit",
        "Contribution Margin %",
        "",
        "Revenue per Unit (net of returns)",
        "Cost per Unit (fully loaded)",
        "Profit per Unit",
    ]
    for label in pl_rows:
        if label == "":
            r += 1
            continue
        data_cell(ws, r, 1, label, font=BOLD_FONT)
        for c in range(2, 7):
            fmt = FMT_PCT if "%" in label else FMT_USD
            yellow(ws, r, c, fmt=fmt)
        r += 1
    r += 1

    # SECTION D — Strategic analysis
    section_label(ws, r, 1, "SECTION D — Channel Strategy Questions  (green = written answer)", 7)
    r += 1
    questions = [
        "Q1: Rank the 4 channels by contribution margin %. Which channel looks best on paper? Does that match where Kevin is investing growth dollars?",
        "Q2: Costco drives volume but at 35% wholesale discount + 28% return rate. Calculate the fully-loaded profit per Costco unit vs. DTC unit. Is Costco worth it?",
        "Q3: Amazon's 15% platform fee + 18% ad spend = 33% of revenue gone before COGS. At what return rate does Amazon become unprofitable? Where is it today?",
        "Q4: B2B / Clinic has the best unit economics by far but is only 4% of volume. Kevin just hired a B2B sales rep. How aggressively would you shift resources here, and what's the risk?",
        "Q5: If you could move 5,000 units from the worst-margin channel to the best-margin channel, what's the P&L impact? Is that realistic?",
        "Q6: Kevin says 'we're a DTC brand.' The data might say something different. How do you present this to Kevin without threatening his identity?",
    ]
    for q in questions:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
        data_cell(ws, r, 1, q, font=BODY_FONT)
        ws.cell(row=r, column=1).alignment = WRAP
        r += 1
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
        green(ws, r, 1)
        ws.row_dimensions[r].height = 55
        r += 1

    ws.sheet_properties.tabColor = "2980B9"


# ═══════════════════════════════════════════════════════════════════════════════

def main():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    build_start(wb)
    build_scenario_1(wb)
    build_scenario_2(wb)
    build_scenario_3(wb)
    build_scenario_4(wb)
    build_scenario_5(wb)
    build_scenario_6(wb)

    path = "/Users/treysisson/iRestore/iRestore_Demand_Planning_Practice.xlsx"
    wb.save(path)
    print(f"✓ Built {path}")
    print(f"  7 sheets (Start Here + 6 scenarios)")


if __name__ == "__main__":
    main()
